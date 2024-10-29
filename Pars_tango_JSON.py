import openpyxl
import json
from tkinter import Tk
from tkinter import filedialog
from openpyxl.styles import Alignment
from datetime import datetime


def main():
    global my_file
    old_file_flag = False

    root = Tk()
    root.title('Оберіть файл')

    root.filename = filedialog.askopenfilename(title="Оберіть файл *.json", filetypes= (("json files", "*.json"), ("all files", "*.*")))
    my_file = root.filename
    print(my_file)
    root.destroy()

    with open(my_file, 'r', encoding='utf-8') as file:
        data = json.load(file)
    row = 2

    try:
        root.filename = filedialog.askopenfilename(title="Оберіть файл *.xlsx", filetypes= (("Excel files", "*.xlsx"), ("all files", "*.*")))
        my_data_file = root.filename
        print(my_data_file)

        book = openpyxl.load_workbook(my_data_file)
        sheet = book.active
        row = sheet.max_row + 1
        if row < 2:
            row = 2
        old_file_flag = True

    except:
        book = openpyxl.Workbook()
        sheet = book.active

    date_file = my_file.split('/')[-1].split('_').pop(1)
    time_h_file = my_file.split('/')[-1].split('_')[2].split('.')[0]
    time_m_file = my_file.split('/')[-1].split('_')[2].split('.')[1]
    date_format = datetime.strptime(date_file, "%d.%m.%Y")
    sheet['A1'] = 'num'
    sheet['B1'] = 'ID стрімера'
    sheet['C1'] = 'ID дарувальника'
    sheet['D1'] = 'Ім`я дарувальника'
    sheet['E1'] = 'Кількість монет'
    sheet['F1'] = 'Посилання на дарувальника'
    sheet['G1'] = 'Ім`я стрімера'
    sheet['H1'] = 'Посилання на стрім'
    sheet['I1'] = 'Посилання на стрімера'
    sheet['J1'] = 'VIP status'
    sheet['K1'] = 'Гендер'
    sheet['L1'] = 'Підписка на стрімера'
    sheet['M1'] = 'Фан рівень'
    sheet['N1'] = 'Incognito'
    sheet['O1'] = 'Дата трансляції'
    sheet['P1'] = 'Час трансляції'

    sheet.column_dimensions["A"].width = 6 # прим. колво символов
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 29
    sheet.column_dimensions["D"].width = 35
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 30
    sheet.column_dimensions["G"].width = 35
    sheet.column_dimensions["H"].width = 30
    sheet.column_dimensions["I"].width = 30
    sheet.column_dimensions["J"].width = 10
    sheet.column_dimensions["K"].width = 17
    sheet.column_dimensions["L"].width = 17
    sheet.column_dimensions["M"].width = 17
    sheet.column_dimensions["N"].width = 17
    sheet.column_dimensions["O"].width = 17
    sheet.column_dimensions["P"].width = 14

    for gifter in data:
        sheet[row][0].value = gifter['num']
        sheet[row][1].value = gifter['ID стрімера']
        sheet[row][2].value = gifter['ID дарувальника']
        sheet[row][3].value = gifter['Ім`я дарувальника']
        sheet[row][4].value = int(gifter['Кількість монет'])
        sheet[row][5].hyperlink = gifter['Посилання дарувальника']
        sheet[row][6].value = gifter['Ім`я стрімера']
        sheet[row][7].hyperlink = gifter['Посилання на стрім']
        sheet[row][8].hyperlink = gifter['Посилання на стрімера']
        sheet[row][9].value = gifter['VIP status']
        sheet[row][10].value = gifter['Гендер']
        sheet[row][11].value = gifter['Підписка на стрімера']
        sheet[row][12].value = gifter['Фан рівень']
        sheet[row][13].value = gifter['Incognito']
        sheet[row][14].value = date_format
        sheet[row][14].number_format = 'DD.mm.YYYY'
        sheet[row][15].value = f"{time_h_file}:{time_m_file}"

        row += 1

    currentcell = sheet['D1']
    currentcell.alignment = Alignment(horizontal='center', wrap_text=True)
    currentcell = sheet['E1']
    currentcell.alignment = Alignment(horizontal='center', wrap_text=True)

    sheet.auto_filter.ref = "A1:P9999"

    if old_file_flag is False:
        root = Tk()
        root.title('Оберіть файл')

        my_file = filedialog.askdirectory(title='Оберіть директорію для збереження файлу EXCEL')
        if my_file == "":
            sys.exit(1)

        print(my_file)
        root.destroy()
        book.save(f'{my_file}/Gifters.xlsx')
    else:
        book.save(my_data_file)
    book.close()


if __name__ == "__main__":
    main()